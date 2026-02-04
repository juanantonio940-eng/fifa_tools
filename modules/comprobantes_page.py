"""
Interfaz web para el Verificador de comprobantes del Mundial FIFA 2026.
Ejecutar con: streamlit run verificador_web.py

Mejoras incluidas:
- OCR gratuito (EasyOCR) con fallback a Claude Vision
- Cache de resultados para ahorrar tiempo y dinero
- Procesamiento en paralelo
- Tolerancia en comparacion de emails
- Vista previa para debug
- Reintentos automaticos si falla Claude Vision
- Exportacion a Excel con formato
"""

import configparser
import shutil
import base64
import json
import time
import hashlib
import re
import os
import ssl
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher
from io import BytesIO
from dotenv import load_dotenv

# Desactivar verificacion SSL para descarga de modelos EasyOCR
ssl._create_default_https_context = ssl._create_unverified_context

import anthropic

# Cargar variables de entorno
env_path = Path(__file__).parent.parent / '.env'
if env_path.exists():
    load_dotenv(env_path)

# ============== OCR (EasyOCR) ==============

# Variable global para el reader de EasyOCR (lazy loading)
_easyocr_reader = None


def get_easyocr_reader():
    """Obtiene el reader de EasyOCR (lazy loading)"""
    global _easyocr_reader
    if _easyocr_reader is None:
        import easyocr
        _easyocr_reader = easyocr.Reader(['en'], gpu=False, verbose=False)
    return _easyocr_reader


def extraer_datos_con_ocr(ruta_imagen: str, motor_ocr: str = "easyocr") -> dict:
    """Extrae datos de una imagen usando EasyOCR (gratuito)"""
    # Solo EasyOCR disponible
    try:
        reader = get_easyocr_reader()
        resultados = reader.readtext(ruta_imagen)

        # Extraer todo el texto
        textos = [r[1] for r in resultados]
        texto_completo = ' '.join(textos).lower()

        # Buscar email (patron comun)
        email = None
        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        emails_encontrados = re.findall(email_pattern, ' '.join(textos))
        if emails_encontrados:
            # Filtrar emails que no sean del sistema (ej: noreply, support, etc)
            for e in emails_encontrados:
                e_lower = e.lower()
                if not any(x in e_lower for x in ['noreply', 'support', 'fifa', 'ticket']):
                    email = e_lower
                    break
            if not email and emails_encontrados:
                email = emails_encontrados[0].lower()

        # Buscar Match number
        match_num = None
        match_patterns = [
            r'match\s*(\d{1,3})',
            r'match\s*#\s*(\d{1,3})',
            r'partido\s*(\d{1,3})'
        ]
        for pattern in match_patterns:
            match_found = re.search(pattern, texto_completo)
            if match_found:
                match_num = int(match_found.group(1))
                break

        # Buscar cantidad de tickets
        cantidad = None
        cantidad_patterns = [
            r'(\d+)\s*tickets?\s*selected',
            r'(\d+)\s*tickets?',
            r'transfer\s*(\d+)',
            r'(\d+)\s*entradas?'
        ]
        for pattern in cantidad_patterns:
            match_found = re.search(pattern, texto_completo)
            if match_found:
                num = int(match_found.group(1))
                if 1 <= num <= 20:  # Rango razonable de tickets
                    cantidad = num
                    break

        # Buscar categoria
        categoria = None
        cat_patterns = [
            r'category\s*(\d+)',
            r'cat\.?\s*(\d+)',
            r'categoria\s*(\d+)'
        ]
        for pattern in cat_patterns:
            match_found = re.search(pattern, texto_completo)
            if match_found:
                categoria = f"Category {match_found.group(1)}"
                break

        return {
            'email': email,
            'match': match_num,
            'cantidad': cantidad,
            'categoria': categoria,
            'metodo': 'OCR',
            'texto_raw': texto_completo[:500],  # Para debug
            'cache': False
        }

    except Exception as e:
        return {
            'email': None,
            'match': None,
            'cantidad': None,
            'categoria': None,
            'metodo': 'OCR',
            'error': str(e),
            'cache': False
        }


def validar_resultado_ocr(resultado: dict) -> tuple:
    """
    Valida si el resultado del OCR es confiable.
    Retorna (es_valido, campos_validos, campos_faltantes)
    """
    campos_validos = []
    campos_faltantes = []

    # Validar email
    if resultado.get('email'):
        email = resultado['email']
        if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            campos_validos.append('email')
        else:
            campos_faltantes.append('email')
    else:
        campos_faltantes.append('email')

    # Validar match (debe ser numero entre 1 y 100)
    if resultado.get('match'):
        if 1 <= resultado['match'] <= 100:
            campos_validos.append('match')
        else:
            campos_faltantes.append('match')
    else:
        campos_faltantes.append('match')

    # Validar cantidad (debe ser numero entre 1 y 20)
    if resultado.get('cantidad'):
        if 1 <= resultado['cantidad'] <= 20:
            campos_validos.append('cantidad')
        else:
            campos_faltantes.append('cantidad')
    else:
        campos_faltantes.append('cantidad')

    # Validar categoria
    if resultado.get('categoria'):
        campos_validos.append('categoria')
    else:
        campos_faltantes.append('categoria')

    # El resultado es valido si tenemos al menos email y match
    es_valido = 'email' in campos_validos and 'match' in campos_validos

    return es_valido, campos_validos, campos_faltantes


# ============== CONFIGURACION POR USUARIO ==============

# Carpeta base para datos de usuarios (configurable)
USUARIOS_BASE_PATH = Path(__file__).parent.parent / 'datos_usuarios'


def get_usuario_email() -> str:
    """Obtiene el email del usuario autenticado de Clerk"""
    import streamlit as st
    user = st.session_state.get('clerk_user', {})
    email = user.get('email', '') if user else ''
    return email.lower().strip() if email else ''


def sanitizar_email_para_carpeta(email: str) -> str:
    """Convierte email en nombre de carpeta seguro"""
    if not email:
        return 'anonimo'
    # Reemplazar caracteres no válidos
    return email.replace('@', '_at_').replace('.', '_').replace(' ', '_')


def get_carpeta_usuario(email: str = None) -> Path:
    """
    Obtiene o crea la carpeta del usuario.
    Estructura: datos_usuarios/<email_sanitizado>/
    """
    if not email:
        email = get_usuario_email()

    carpeta_nombre = sanitizar_email_para_carpeta(email)
    carpeta_usuario = USUARIOS_BASE_PATH / carpeta_nombre

    # Crear estructura de carpetas si no existe
    if not carpeta_usuario.exists():
        carpeta_usuario.mkdir(parents=True, exist_ok=True)
        (carpeta_usuario / 'imagenes').mkdir(exist_ok=True)
        (carpeta_usuario / 'tabla').mkdir(exist_ok=True)
        (carpeta_usuario / 'reportes').mkdir(exist_ok=True)
        print(f"[Comprobantes] Carpeta creada para usuario: {email}")

    return carpeta_usuario


def get_config_path(email: str = None):
    """Retorna la ruta del archivo de configuracion del usuario"""
    carpeta = get_carpeta_usuario(email)
    return carpeta / 'config.ini'


def get_cache_path(email: str = None):
    """Retorna la ruta del archivo de cache del usuario"""
    carpeta = get_carpeta_usuario(email)
    return carpeta / 'cache_resultados.json'


def cargar_config() -> dict:
    """Carga la configuracion del usuario desde su carpeta"""
    email = get_usuario_email()
    carpeta_usuario = get_carpeta_usuario(email)

    config = configparser.ConfigParser()
    config_path = get_config_path(email)

    if not config_path.exists():
        crear_config_defecto(email)

    config.read(config_path)

    # API key: primero del .env, luego del config.ini
    api_key = os.getenv('ANTHROPIC_API_KEY', '')
    if not api_key:
        api_key = config['api'].get('anthropic_api_key', '')

    return {
        'imagenes': Path(config['rutas']['imagenes']),
        'tabla': Path(config['rutas']['tabla']),
        'reporte': Path(config['rutas']['reporte']),
        'api_key': api_key,
        'config_path': config_path
    }


def crear_config_defecto(email: str = None):
    """Crea un archivo de configuracion por defecto para el usuario"""
    carpeta_usuario = get_carpeta_usuario(email)

    config = configparser.ConfigParser()
    config['rutas'] = {
        'imagenes': str(carpeta_usuario / 'imagenes'),
        'tabla': str(carpeta_usuario / 'tabla' / 'datos_referencia.csv'),
        'reporte': str(carpeta_usuario / 'reportes')
    }
    config['api'] = {
        'anthropic_api_key': ''
    }
    with open(get_config_path(email), 'w') as f:
        config.write(f)


def explorador_carpetas(ruta_actual: str, key_prefix: str, modo: str = "carpeta") -> str:
    """
    Muestra un explorador de carpetas/archivos.
    modo: "carpeta" para seleccionar carpetas, "archivo" para archivos
    Retorna la ruta seleccionada o la ruta actual si no se selecciona nada.
    """
    try:
        ruta = Path(ruta_actual) if ruta_actual else Path.home()
        if not ruta.exists():
            ruta = Path.home()

        # Obtener carpeta padre
        if ruta.is_file():
            carpeta_actual = ruta.parent
        else:
            carpeta_actual = ruta

        # Inicializar estado
        state_key = f"explorer_{key_prefix}"
        if state_key not in st.session_state:
            st.session_state[state_key] = str(carpeta_actual)

        carpeta_nav = Path(st.session_state[state_key])

        # Mostrar ruta actual
        st.caption(f"📂 {carpeta_nav}")

        # Botones de navegación
        col1, col2 = st.columns([1, 4])
        with col1:
            if st.button("⬆️ Subir", key=f"up_{key_prefix}"):
                if carpeta_nav.parent != carpeta_nav:
                    st.session_state[state_key] = str(carpeta_nav.parent)
                    st.rerun()

        # Listar contenido
        try:
            items = sorted(carpeta_nav.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower()))

            # Filtrar según modo
            if modo == "carpeta":
                items = [i for i in items if i.is_dir()]
            elif modo == "archivo":
                items = [i for i in items if i.is_dir() or i.suffix.lower() in ['.csv', '.xlsx', '.xls']]

            if items:
                opciones = ["-- Seleccionar --"] + [f"{'📁' if i.is_dir() else '📄'} {i.name}" for i in items]
                seleccion = st.selectbox(
                    "Navegar:",
                    opciones,
                    key=f"nav_{key_prefix}",
                    label_visibility="collapsed"
                )

                if seleccion != "-- Seleccionar --":
                    nombre = seleccion[2:].strip()  # Quitar emoji
                    item_seleccionado = carpeta_nav / nombre

                    if item_seleccionado.is_dir():
                        if modo == "carpeta":
                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button("📂 Entrar", key=f"enter_{key_prefix}"):
                                    st.session_state[state_key] = str(item_seleccionado)
                                    st.rerun()
                            with col2:
                                if st.button("✅ Seleccionar esta", key=f"sel_{key_prefix}"):
                                    return str(item_seleccionado)
                        else:
                            if st.button("📂 Entrar", key=f"enter_{key_prefix}"):
                                st.session_state[state_key] = str(item_seleccionado)
                                st.rerun()
                    else:
                        if st.button("✅ Seleccionar archivo", key=f"sel_{key_prefix}"):
                            return str(item_seleccionado)
            else:
                st.info("Carpeta vacía o sin elementos válidos")

        except PermissionError:
            st.error("Sin permisos para acceder a esta carpeta")

        # Opción de seleccionar carpeta actual
        if modo == "carpeta":
            if st.button("✅ Usar carpeta actual", key=f"use_current_{key_prefix}"):
                return str(carpeta_nav)

    except Exception as e:
        st.error(f"Error: {e}")

    return ruta_actual


def guardar_config(imagenes: str, tabla: str, reporte: str, api_key: str, email: str = None):
    """Guarda la configuracion del usuario en su carpeta"""
    if not email:
        email = get_usuario_email()

    config = configparser.ConfigParser()
    config['rutas'] = {
        'imagenes': imagenes,
        'tabla': tabla,
        'reporte': reporte
    }
    config['api'] = {
        'anthropic_api_key': api_key
    }
    with open(get_config_path(email), 'w') as f:
        config.write(f)


# ============== CACHE ==============

def cargar_cache() -> dict:
    """Carga el cache de resultados"""
    cache_path = get_cache_path()
    if cache_path.exists():
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def guardar_cache(cache: dict):
    """Guarda el cache de resultados"""
    with open(get_cache_path(), 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def calcular_hash_imagen(ruta_imagen: str) -> str:
    """Calcula el hash MD5 de una imagen para identificarla en cache"""
    with open(ruta_imagen, 'rb') as f:
        return hashlib.md5(f.read()).hexdigest()


def limpiar_cache():
    """Elimina el archivo de cache"""
    cache_path = get_cache_path()
    if cache_path.exists():
        cache_path.unlink()


# ============== DATOS ==============

def cargar_tabla(ruta: str) -> pd.DataFrame:
    """Carga la tabla de referencia."""
    ruta = Path(ruta)
    if ruta.suffix.lower() == '.csv':
        # Intentar diferentes codificaciones y separadores
        separadores = [',', ';', '\t', '|']
        codificaciones = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']

        for sep in separadores:
            for encoding in codificaciones:
                try:
                    df = pd.read_csv(ruta, encoding=encoding, sep=sep)
                    # Verificar que tiene mas de una columna (separador correcto)
                    if len(df.columns) > 1:
                        print(f"[CSV] Cargado con sep='{sep}', encoding={encoding}")
                        return df
                except:
                    continue

        # Fallback: intentar con deteccion automatica
        return pd.read_csv(ruta, encoding='utf-8', errors='ignore')
    elif ruta.suffix.lower() in ['.xlsx', '.xls']:
        return pd.read_excel(ruta)
    else:
        raise ValueError(f"Formato no soportado: {ruta.suffix}")


def agrupar_por_pedido(tabla: pd.DataFrame) -> dict:
    """Agrupa los datos por PEDIDO VENTA."""
    agrupados = {}

    # Verificar columnas requeridas
    columnas_requeridas = ['PEDIDO VENTA', 'match', 'email_envio']
    columnas_tabla = list(tabla.columns)

    for col in columnas_requeridas:
        if col not in columnas_tabla:
            print(f"[Error] Columna '{col}' no encontrada. Columnas disponibles: {columnas_tabla}")
            raise KeyError(f"Columna '{col}' no encontrada en la tabla. Columnas disponibles: {columnas_tabla}")

    for pedido in tabla['PEDIDO VENTA'].unique():
        filas = tabla[tabla['PEDIDO VENTA'] == pedido]
        primera = filas.iloc[0]

        if pd.notna(pedido):
            try:
                pedido_str = str(int(float(pedido)))
            except (ValueError, TypeError):
                pedido_str = str(pedido).strip()
        else:
            continue

        # Obtener campos con valores por defecto si no existen
        match_val = primera.get('match', 0)
        team_a = primera.get('Team A', 'N/A')
        team_b = primera.get('Team B', 'N/A')
        categoria = primera.get('Category', 'N/A')
        email = primera.get('email_envio', 'N/A')

        agrupados[pedido_str] = {
            'match': int(match_val) if pd.notna(match_val) else 0,
            'partido': f"{team_a} vs {team_b}",
            'categoria': categoria,
            'cantidad': len(filas),
            'email': str(email).strip().lower() if pd.notna(email) else ''
        }

    return agrupados


def obtener_imagenes(carpeta: Path) -> list:
    """Obtiene las imagenes de la carpeta."""
    extensiones = {'.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG'}
    return [f for f in carpeta.iterdir() if f.suffix in extensiones]


# ============== TOLERANCIA EMAIL ==============

def similitud_email(email1: str, email2: str) -> float:
    """Calcula la similitud entre dos emails (0-1)"""
    if not email1 or not email2:
        return 0.0
    return SequenceMatcher(None, email1.lower(), email2.lower()).ratio()


def comparar_emails(email_img: str, email_tabla: str, tolerancia: float = 0.9) -> tuple:
    """
    Compara dos emails con tolerancia.
    Retorna (coincide_exacto, coincide_con_tolerancia, similitud)
    """
    if not email_img or not email_tabla:
        return False, False, 0.0

    email_img = email_img.lower().strip()
    email_tabla = email_tabla.lower().strip()

    # Coincidencia exacta
    if email_img == email_tabla:
        return True, True, 1.0

    # Calcular similitud
    similitud = similitud_email(email_img, email_tabla)

    # Verificar partes del email por separado
    try:
        user1, domain1 = email_img.split('@')
        user2, domain2 = email_tabla.split('@')

        # Si el dominio es igual, dar mas peso
        if domain1 == domain2:
            user_sim = SequenceMatcher(None, user1, user2).ratio()
            similitud = max(similitud, 0.5 + (user_sim * 0.5))
    except:
        pass

    return False, similitud >= tolerancia, similitud


# ============== CLAUDE VISION ==============

def extraer_datos_con_claude(ruta_imagen: str, client, max_reintentos: int = 3, delay: float = 2.0) -> dict:
    """Extrae datos de una imagen usando Claude Vision con reintentos automaticos."""

    with open(ruta_imagen, 'rb') as f:
        imagen_bytes = f.read()

    imagen_base64 = base64.standard_b64encode(imagen_bytes).decode('utf-8')

    extension = Path(ruta_imagen).suffix.lower()
    media_type = "image/jpeg" if extension in ['.jpg', '.jpeg'] else "image/png"

    prompt = """Analiza esta imagen de un comprobante de tickets del Mundial FIFA 2026.

Extrae los siguientes datos que estan visibles en la imagen:
1. **email**: El email que aparece en el campo "Transfer Recipient's email address" (dentro del modal/formulario)
2. **match**: El numero de partido que aparece en la parte SUPERIOR de la imagen (formato "Match X" o "Match XX")
3. **cantidad**: La cantidad de tickets que se estan TRANSFIRIENDO. IMPORTANTE: Busca "X tickets" en la parte INFERIOR de la imagen (en la barra de abajo, cerca del boton "TRANSFER TICKET(S)"). NO uses el numero de arriba que muestra el total de tickets disponibles.
4. **categoria**: La categoria que aparece junto a los tickets (formato "Category X")

IMPORTANTE:
- El match SIEMPRE esta en la parte superior de la imagen, no uses el que pueda aparecer detras del modal
- El email esta dentro del campo de texto del formulario
- La cantidad de tickets a transferir esta en la parte INFERIOR de la pantalla (barra de abajo), NO en la parte superior. La parte superior muestra el total, pero la inferior muestra los seleccionados con checkbox.

Responde SOLO con un JSON valido en este formato exacto:
{"email": "email@ejemplo.com", "match": 25, "cantidad": 4, "categoria": 3}

Si no puedes leer algun campo, usa null para ese campo."""

    ultimo_error = None

    for intento in range(max_reintentos):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=200,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": media_type,
                                    "data": imagen_base64
                                }
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ]
            )

            respuesta_texto = response.content[0].text.strip()

            if respuesta_texto.startswith('```'):
                respuesta_texto = respuesta_texto.split('```')[1]
                if respuesta_texto.startswith('json'):
                    respuesta_texto = respuesta_texto[4:]
                respuesta_texto = respuesta_texto.strip()

            datos = json.loads(respuesta_texto)

            return {
                'email': datos.get('email', '').lower() if datos.get('email') else None,
                'match': datos.get('match'),
                'cantidad': datos.get('cantidad'),
                'categoria': f"Category {datos.get('categoria')}" if datos.get('categoria') else None,
                'metodo': 'Claude Vision',
                'reintentos': intento,
                'cache': False
            }

        except Exception as e:
            ultimo_error = str(e)
            if intento < max_reintentos - 1:
                time.sleep(delay * (intento + 1))

    return {
        'email': None,
        'match': None,
        'cantidad': None,
        'categoria': None,
        'metodo': 'Claude Vision',
        'error': f"Fallo despues de {max_reintentos} intentos: {ultimo_error}",
        'reintentos': max_reintentos,
        'cache': False
    }


# ============== EXTRACCION INTELIGENTE ==============

def extraer_datos_inteligente(ruta_imagen: str, client, metodo: str = "ocr_fallback", motor_ocr: str = "easyocr") -> dict:
    """
    Extrae datos usando el metodo especificado:
    - "solo_ocr": Solo usa OCR (gratuito)
    - "solo_claude": Solo usa Claude Vision (de pago)
    - "ocr_fallback": Intenta OCR primero, si falla usa Claude Vision

    motor_ocr puede ser "easyocr" o "paddleocr"
    """

    if metodo == "solo_claude":
        return extraer_datos_con_claude(ruta_imagen, client)

    elif metodo == "solo_ocr":
        return extraer_datos_con_ocr(ruta_imagen, motor_ocr)

    else:  # ocr_fallback (por defecto)
        # Primero intentar con OCR
        resultado_ocr = extraer_datos_con_ocr(ruta_imagen, motor_ocr)

        # Validar resultado
        es_valido, campos_validos, campos_faltantes = validar_resultado_ocr(resultado_ocr)

        if es_valido:
            # OCR fue suficiente
            resultado_ocr['fallback_usado'] = False
            return resultado_ocr
        else:
            # Necesitamos Claude Vision
            resultado_claude = extraer_datos_con_claude(ruta_imagen, client)
            resultado_claude['fallback_usado'] = True
            resultado_claude['ocr_campos_encontrados'] = campos_validos
            resultado_claude['metodo'] = 'Claude Vision (fallback)'
            return resultado_claude


def extraer_con_cache(ruta_imagen: str, client, cache: dict, usar_cache: bool = True, metodo: str = "ocr_fallback", motor_ocr: str = "easyocr") -> dict:
    """Extrae datos usando cache si esta disponible"""
    imagen_hash = calcular_hash_imagen(ruta_imagen)

    if usar_cache and imagen_hash in cache:
        resultado = cache[imagen_hash].copy()
        resultado['cache'] = True
        return resultado

    resultado = extraer_datos_inteligente(ruta_imagen, client, metodo, motor_ocr)

    # Guardar en cache si no hay error
    if 'error' not in resultado:
        cache[imagen_hash] = {
            'email': resultado.get('email'),
            'match': resultado.get('match'),
            'cantidad': resultado.get('cantidad'),
            'categoria': resultado.get('categoria'),
            'metodo': resultado.get('metodo'),
            'fecha_cache': datetime.now().isoformat()
        }

    return resultado


# ============== PROCESAMIENTO ==============

def procesar_imagen_worker(args):
    """Worker para procesamiento paralelo"""
    imagen, client, cache, usar_cache, datos_tabla, tolerancia_email, metodo_extraccion, motor_ocr = args

    pedido = imagen.stem
    resultado = {
        'archivo': imagen.name,
        'pedido': pedido
    }

    # Extraer datos
    datos_img = extraer_con_cache(str(imagen), client, cache, usar_cache, metodo_extraccion, motor_ocr)

    resultado['email_imagen'] = datos_img.get('email')
    resultado['match_imagen'] = datos_img.get('match')
    resultado['cantidad_imagen'] = datos_img.get('cantidad')
    resultado['categoria_imagen'] = datos_img.get('categoria')
    resultado['desde_cache'] = datos_img.get('cache', False)
    resultado['metodo_usado'] = datos_img.get('metodo', 'desconocido')
    resultado['fallback_usado'] = datos_img.get('fallback_usado', False)
    resultado['reintentos'] = datos_img.get('reintentos', 0)
    resultado['error_extraccion'] = datos_img.get('error')

    # Buscar en tabla
    if pedido not in datos_tabla:
        resultado['estado'] = 'NO ENCONTRADO EN TABLA'
        resultado['email_tabla'] = ''
        resultado['match_tabla'] = ''
        resultado['cantidad_tabla'] = ''
        resultado['categoria_tabla'] = ''
        resultado['email_ok'] = False
        resultado['email_similar'] = False
        resultado['similitud_email'] = 0.0
        resultado['match_ok'] = False
        resultado['cantidad_ok'] = False
        resultado['categoria_ok'] = False
        return resultado

    info_tabla = datos_tabla[pedido]
    resultado['email_tabla'] = info_tabla['email']
    resultado['match_tabla'] = info_tabla['match']
    resultado['cantidad_tabla'] = info_tabla['cantidad']
    resultado['categoria_tabla'] = info_tabla['categoria']

    # Comparar email con tolerancia
    email_exacto, email_tolerancia, similitud = comparar_emails(
        datos_img.get('email'),
        info_tabla['email'],
        tolerancia_email
    )
    resultado['email_ok'] = email_exacto
    resultado['email_similar'] = email_tolerancia and not email_exacto
    resultado['similitud_email'] = round(similitud * 100, 1)

    # Comparar otros campos
    resultado['match_ok'] = datos_img.get('match') == info_tabla['match'] if datos_img.get('match') else False
    resultado['cantidad_ok'] = datos_img.get('cantidad') == info_tabla['cantidad'] if datos_img.get('cantidad') else False

    cat_img_num = ''.join(filter(str.isdigit, datos_img.get('categoria', '') or ''))
    cat_tabla_num = ''.join(filter(str.isdigit, str(info_tabla['categoria'])))
    resultado['categoria_ok'] = cat_img_num == cat_tabla_num if cat_img_num else False

    # Determinar estado
    email_valido = email_exacto or email_tolerancia
    if email_valido and resultado['match_ok'] and resultado['cantidad_ok'] and resultado['categoria_ok']:
        if email_exacto:
            resultado['estado'] = "OK"
        else:
            resultado['estado'] = "OK (EMAIL SIMILAR)"
    elif email_valido:
        resultado['estado'] = "PARCIAL"
    else:
        resultado['estado'] = "NO COINCIDE"

    return resultado


# ============== CLASIFICACION ==============

def clasificar_archivos(resultados, carpeta_imagenes, carpeta_reporte):
    """Clasifica los archivos en carpetas"""
    carpeta_buenos = carpeta_reporte / 'buenos'
    carpeta_regular = carpeta_reporte / 'regular'
    carpeta_malos = carpeta_reporte / 'malos'

    for carpeta in [carpeta_buenos, carpeta_regular, carpeta_malos]:
        if carpeta.exists():
            shutil.rmtree(carpeta)
        carpeta.mkdir(exist_ok=True)

    for r in resultados:
        archivo_origen = carpeta_imagenes / r['archivo']

        if r['estado'] in ['OK', 'OK (EMAIL SIMILAR)']:
            destino = carpeta_buenos / r['archivo']
        elif r['estado'] == 'PARCIAL':
            destino = carpeta_regular / r['archivo']
        else:
            destino = carpeta_malos / r['archivo']

        shutil.copy2(archivo_origen, destino)

    return carpeta_buenos, carpeta_regular, carpeta_malos


# ============== EXPORTACION EXCEL ==============

def exportar_excel(df: pd.DataFrame) -> bytes:
    """Exporta el DataFrame a Excel con formato"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Verificacion"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    parcial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Escribir datos
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if r_idx == 1:  # Header
                cell.font = header_font
                cell.fill = header_fill
            else:
                # Colorear segun estado
                if c_idx == 2:  # Columna estado
                    if value in ['OK', 'OK (EMAIL SIMILAR)']:
                        cell.fill = ok_fill
                    elif value == 'PARCIAL':
                        cell.fill = parcial_fill
                    else:
                        cell.fill = error_fill

    # Ajustar anchos
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Crear hoja de resumen
    ws_resumen = wb.create_sheet(title="Resumen")

    # Contar estados
    conteo = df['estado'].value_counts()

    ws_resumen['A1'] = "RESUMEN DE VERIFICACION"
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A3'] = "Estado"
    ws_resumen['B3'] = "Cantidad"
    ws_resumen['A3'].font = header_font
    ws_resumen['B3'].font = header_font
    ws_resumen['A3'].fill = header_fill
    ws_resumen['B3'].fill = header_fill

    row = 4
    for estado, cantidad in conteo.items():
        ws_resumen[f'A{row}'] = estado
        ws_resumen[f'B{row}'] = cantidad
        row += 1

    ws_resumen[f'A{row}'] = "TOTAL"
    ws_resumen[f'B{row}'] = len(df)
    ws_resumen[f'A{row}'].font = Font(bold=True)
    ws_resumen[f'B{row}'].font = Font(bold=True)

    # Estadisticas de metodo
    if 'metodo_usado' in df.columns:
        row += 3
        ws_resumen[f'A{row}'] = "METODO DE EXTRACCION"
        ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        metodo_conteo = df['metodo_usado'].value_counts()
        for metodo, cant in metodo_conteo.items():
            ws_resumen[f'A{row}'] = metodo
            ws_resumen[f'B{row}'] = cant
            row += 1

    ws_resumen['A{}'.format(row + 2)] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Guardar en bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============== INTERFAZ PRINCIPAL ==============

def render():
    """Renderiza la página de verificación de comprobantes"""
    st.title("⚽ Verificador de Comprobantes - Mundial FIFA 2026")

    # Tabs principales
    tab_verificacion, tab_config, tab_debug = st.tabs(["🔍 Verificacion", "⚙️ Configuracion", "🔧 Debug/Preview"])

    # === TAB CONFIGURACION ===
    with tab_config:
        st.header("⚙️ Configuracion")

        # Obtener datos del usuario
        email_usuario = get_usuario_email()
        carpeta_usuario = get_carpeta_usuario(email_usuario)
        config_actual = cargar_config()

        # Mostrar info del usuario
        st.info(f"👤 Usuario: **{email_usuario or 'No autenticado'}**")
        st.caption(f"📂 Tu carpeta de datos: `{carpeta_usuario}`")

        st.markdown("---")

        # === SUBIR IMAGENES ===
        st.subheader("📷 Subir Imagenes de Comprobantes")

        imagenes_subidas = st.file_uploader(
            "Arrastra o selecciona imagenes de comprobantes",
            type=['jpg', 'jpeg', 'png'],
            accept_multiple_files=True,
            help="Sube las capturas de pantalla de los comprobantes de transferencia",
            key="upload_imagenes"
        )

        if imagenes_subidas:
            # Mostrar archivos seleccionados
            st.info(f"📁 {len(imagenes_subidas)} archivos seleccionados")

            if st.button("💾 Guardar imagenes", key="btn_guardar_imgs", type="primary"):
                carpeta_imagenes = carpeta_usuario / 'imagenes'
                guardadas = 0
                for img_file in imagenes_subidas:
                    ruta_destino = carpeta_imagenes / img_file.name
                    with open(ruta_destino, 'wb') as f:
                        f.write(img_file.getbuffer())
                    guardadas += 1
                    print(f"[Guardado] {img_file.name}")
                st.success(f"✅ {guardadas} imagenes guardadas")
                st.rerun()

        # Mostrar imagenes existentes
        carpeta_imagenes = carpeta_usuario / 'imagenes'
        if carpeta_imagenes.exists():
            # Usar set para evitar duplicados (Windows es case-insensitive)
            imgs_set = set()
            for ext in ['*.jpg', '*.jpeg', '*.png']:
                for img in carpeta_imagenes.glob(ext):
                    imgs_set.add(img.resolve())  # resolve() normaliza la ruta
            imgs_existentes = list(imgs_set)

            if imgs_existentes:
                col_info, col_btn = st.columns([3, 1])
                with col_info:
                    st.write(f"📷 **{len(imgs_existentes)} imagenes** en tu carpeta")
                with col_btn:
                    # Key dinámico basado en cantidad de imágenes para evitar caché
                    if st.button("🗑️ Borrar todas", key=f"del_imgs_{len(imgs_existentes)}", type="secondary"):
                        borradas = 0
                        for img in imgs_existentes:
                            try:
                                if img.exists():
                                    img.unlink()
                                    borradas += 1
                                    print(f"[Borrado] {img.name}")
                            except Exception as e:
                                print(f"[Error] borrando {img}: {e}")
                        print(f"[Total] {borradas} imagenes eliminadas")
                        st.success(f"✅ {borradas} imagenes eliminadas")
                        st.rerun()

                with st.expander("👁️ Ver imagenes", expanded=False):
                    cols = st.columns(4)
                    for i, img in enumerate(imgs_existentes[:12]):
                        with cols[i % 4]:
                            st.image(str(img), caption=img.name, width=150)
                    if len(imgs_existentes) > 12:
                        st.caption(f"... y {len(imgs_existentes) - 12} mas")

        st.markdown("---")

        # === SUBIR TABLA DE REFERENCIA ===
        st.subheader("📊 Subir Tabla de Referencia")

        tabla_subida = st.file_uploader(
            "Arrastra o selecciona archivo CSV o Excel",
            type=['csv', 'xlsx', 'xls'],
            accept_multiple_files=False,
            help="Archivo con los datos de referencia (PEDIDO VENTA, match, email_envio, etc.)",
            key="upload_tabla"
        )

        if tabla_subida:
            carpeta_tabla = carpeta_usuario / 'tabla'
            ruta_tabla = carpeta_tabla / tabla_subida.name
            with open(ruta_tabla, 'wb') as f:
                f.write(tabla_subida.getbuffer())
            st.success(f"✅ Tabla guardada: {tabla_subida.name}")

            # Actualizar config con la nueva tabla
            guardar_config(
                str(carpeta_usuario / 'imagenes'),
                str(ruta_tabla),
                str(carpeta_usuario / 'reportes'),
                config_actual.get('api_key', '')
            )

        # Mostrar tabla existente
        carpeta_tabla = carpeta_usuario / 'tabla'
        if carpeta_tabla.exists():
            tablas_existentes = list(carpeta_tabla.glob('*.csv')) + \
                               list(carpeta_tabla.glob('*.xlsx')) + \
                               list(carpeta_tabla.glob('*.xls'))
            if tablas_existentes:
                tabla_actual = tablas_existentes[0]  # Usar la primera
                st.success(f"📊 Tabla actual: **{tabla_actual.name}**")

                # Preview de la tabla
                with st.expander("👁️ Vista previa de la tabla", expanded=False):
                    try:
                        if tabla_actual.suffix == '.csv':
                            df_preview = pd.read_csv(tabla_actual)
                        else:
                            df_preview = pd.read_excel(tabla_actual)
                        st.dataframe(df_preview.head(10), use_container_width=True)
                        st.caption(f"Mostrando 10 de {len(df_preview)} filas")
                    except Exception as e:
                        st.error(f"Error leyendo tabla: {e}")

                if st.button("🗑️ Borrar tabla", key="del_tabla"):
                    for t in tablas_existentes:
                        t.unlink()
                    st.success("Tabla eliminada")
                    st.rerun()

        st.markdown("---")

        # === REPORTES ===
        st.subheader("📁 Reportes Generados")

        carpeta_reportes = carpeta_usuario / 'reportes'
        if carpeta_reportes.exists():
            reportes = list(carpeta_reportes.glob('*.xlsx')) + \
                      list(carpeta_reportes.glob('*.csv'))
            if reportes:
                st.write(f"📄 {len(reportes)} reportes disponibles:")
                for rep in sorted(reportes, key=lambda x: x.stat().st_mtime, reverse=True)[:5]:
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.write(f"• {rep.name}")
                    with col2:
                        with open(rep, 'rb') as f:
                            st.download_button(
                                "⬇️",
                                data=f.read(),
                                file_name=rep.name,
                                mime="application/octet-stream",
                                key=f"dl_{rep.name}"
                            )
                if st.button("🗑️ Borrar todos los reportes", key="del_reportes"):
                    for rep in reportes:
                        rep.unlink()
                    st.success("Reportes eliminados")
                    st.rerun()
            else:
                st.info("No hay reportes generados aun")

        st.markdown("---")

        # === API KEY ===
        st.subheader("🔑 API de Anthropic (Claude Vision)")

        api_key_actual = config_actual.get('api_key', '') if config_actual else ''
        nueva_api_key = st.text_input(
            "API Key",
            value=api_key_actual,
            type="password",
            help="Obten tu API key en: https://console.anthropic.com/"
        )

        if nueva_api_key != api_key_actual:
            if st.button("💾 Guardar API Key"):
                guardar_config(
                    str(carpeta_usuario / 'imagenes'),
                    str(config_actual.get('tabla', carpeta_usuario / 'tabla' / 'datos.csv')),
                    str(carpeta_usuario / 'reportes'),
                    nueva_api_key
                )
                st.success("✅ API Key guardada")
                st.rerun()

        st.markdown("---")

        # === RESUMEN ===
        st.subheader("✅ Estado de tu configuracion")

        # Contar imagenes
        num_imgs = len(list((carpeta_usuario / 'imagenes').glob('*.jpg')) +
                      list((carpeta_usuario / 'imagenes').glob('*.jpeg')) +
                      list((carpeta_usuario / 'imagenes').glob('*.png')))

        # Verificar tabla
        tablas = list((carpeta_usuario / 'tabla').glob('*.csv')) + \
                list((carpeta_usuario / 'tabla').glob('*.xlsx'))
        tiene_tabla = len(tablas) > 0

        col1, col2, col3 = st.columns(3)
        with col1:
            if num_imgs > 0:
                st.success(f"📷 {num_imgs} imagenes")
            else:
                st.error("📷 Sin imagenes")
        with col2:
            if tiene_tabla:
                st.success("📊 Tabla OK")
            else:
                st.error("📊 Sin tabla")
        with col3:
            if nueva_api_key:
                st.success("🔑 API OK")
            else:
                st.warning("🔑 Sin API")

    # === TAB DEBUG/PREVIEW ===
    with tab_debug:
        st.header("🔧 Vista Previa y Debug")

        config = cargar_config()

        # Sub-tabs dentro de Debug
        debug_tab1, debug_tab2, debug_tab3, debug_tab4 = st.tabs([
            "🔍 Imagen Individual",
            "🔬 Comparar OCR vs Claude",
            "📊 Analisis Masivo OCR",
            "💾 Cache"
        ])

        # === TAB: Imagen Individual ===
        with debug_tab1:
            st.subheader("🔍 Probar Imagen Individual")

            if config['imagenes'].exists():
                imagenes = obtener_imagenes(config['imagenes'])
                if imagenes:
                    imagen_seleccionada = st.selectbox(
                        "Selecciona una imagen para probar",
                        options=imagenes,
                        format_func=lambda x: x.name,
                        key="debug_img_select"
                    )

                    col1, col2 = st.columns(2)

                    with col1:
                        st.image(str(imagen_seleccionada), caption=imagen_seleccionada.name, width=400)

                    with col2:
                        metodo_test = st.radio(
                            "Metodo de extraccion",
                            ["OCR (Gratuito)", "Claude Vision", "OCR + Fallback"],
                            horizontal=True
                        )

                        metodo_map = {
                            "OCR (Gratuito)": "solo_ocr",
                            "Claude Vision": "solo_claude",
                            "OCR + Fallback": "ocr_fallback"
                        }

                        if st.button("🔬 Analizar Imagen"):
                            with st.spinner(f"Analizando imagen con {metodo_test}..."):
                                metodo = metodo_map[metodo_test]

                                if metodo == "solo_ocr":
                                    resultado = extraer_datos_con_ocr(str(imagen_seleccionada))
                                    es_valido, campos_ok, campos_fail = validar_resultado_ocr(resultado)
                                    st.write(f"**Validacion OCR (EasyOCR):** {'✅ Valido' if es_valido else '❌ Invalido'}")
                                    st.write(f"Campos OK: {campos_ok}")
                                    st.write(f"Campos faltantes: {campos_fail}")
                                elif metodo == "solo_claude":
                                    client = anthropic.Anthropic(api_key=config['api_key'])
                                    resultado = extraer_datos_con_claude(str(imagen_seleccionada), client)
                                else:
                                    client = anthropic.Anthropic(api_key=config['api_key'])
                                    resultado = extraer_datos_inteligente(str(imagen_seleccionada), client, metodo)
                                    if resultado.get('fallback_usado'):
                                        st.warning("⚠️ EasyOCR fallo, se uso Claude Vision como fallback")
                                    else:
                                        st.success("✅ EasyOCR fue suficiente, no se uso Claude Vision")

                                # Mostrar texto raw extraido
                                if resultado.get('texto_raw'):
                                    with st.expander("📝 Texto RAW extraido por OCR"):
                                        st.code(resultado.get('texto_raw', 'No disponible'))

                                st.json(resultado)

                                # Comparar con tabla si existe
                                pedido = imagen_seleccionada.stem
                                if config['tabla'].exists():
                                    tabla = cargar_tabla(str(config['tabla']))
                                    datos_tabla = agrupar_por_pedido(tabla)

                                    if pedido in datos_tabla:
                                        st.subheader("📊 Datos en Tabla")
                                        st.json(datos_tabla[pedido])

                                        # Comparar
                                        st.subheader("🔄 Comparacion")
                                        info = datos_tabla[pedido]

                                        exacto, tolerancia, sim = comparar_emails(resultado.get('email'), info['email'])

                                        comparaciones = {
                                            'Email': f"{'✅' if exacto else ('⚠️' if tolerancia else '❌')} {resultado.get('email')} vs {info['email']} ({sim*100:.0f}%)",
                                            'Match': f"{'✅' if resultado.get('match') == info['match'] else '❌'} {resultado.get('match')} vs {info['match']}",
                                            'Cantidad': f"{'✅' if resultado.get('cantidad') == info['cantidad'] else '❌'} {resultado.get('cantidad')} vs {info['cantidad']}",
                                            'Categoria': f"{resultado.get('categoria')} vs {info['categoria']}"
                                        }

                                        for campo, valor in comparaciones.items():
                                            st.write(f"**{campo}:** {valor}")
                                    else:
                                        st.warning(f"Pedido {pedido} no encontrado en tabla")

        # === TAB: Comparar OCR vs Claude ===
        with debug_tab2:
            st.subheader("🔬 Comparar Motores OCR vs Claude Vision")
            st.info("Compara los diferentes metodos de extraccion en una sola imagen")

            if config['imagenes'].exists():
                imagenes = obtener_imagenes(config['imagenes'])
                if imagenes:
                    imagen_comparar = st.selectbox(
                        "Selecciona imagen para comparar",
                        options=imagenes,
                        format_func=lambda x: x.name,
                        key="debug_compare_select"
                    )

                    st.image(str(imagen_comparar), caption=imagen_comparar.name, width=500)

                    if st.button("🔬 Comparar EasyOCR vs Claude Vision", type="primary"):
                        col1, col2 = st.columns(2)

                        # EasyOCR
                        with col1:
                            st.markdown("### 🆓 EasyOCR (Gratuito)")
                            with st.spinner("Analizando..."):
                                res_easy = extraer_datos_con_ocr(str(imagen_comparar))
                                valido_easy, ok_easy, fail_easy = validar_resultado_ocr(res_easy)

                            if valido_easy:
                                st.success("✅ Valido - No necesita Claude")
                            else:
                                st.error("❌ Invalido - Requiere Claude")

                            st.write(f"**Email:** {res_easy.get('email') or '❌'}")
                            st.write(f"**Match:** {res_easy.get('match') or '❌'}")
                            st.write(f"**Cantidad:** {res_easy.get('cantidad') or '❌'}")
                            st.write(f"**Categoria:** {res_easy.get('categoria') or '❌'}")

                            with st.expander("📝 Texto RAW"):
                                st.code(res_easy.get('texto_raw', 'N/A'))

                        # Claude Vision
                        with col2:
                            st.markdown("### 💰 Claude Vision")
                            with st.spinner("Analizando..."):
                                client = anthropic.Anthropic(api_key=config['api_key'])
                                res_claude = extraer_datos_con_claude(str(imagen_comparar), client)

                            st.info("De pago pero preciso")

                            st.write(f"**Email:** {res_claude.get('email') or '❌'}")
                            st.write(f"**Match:** {res_claude.get('match') or '❌'}")
                            st.write(f"**Cantidad:** {res_claude.get('cantidad') or '❌'}")
                            st.write(f"**Categoria:** {res_claude.get('categoria') or '❌'}")

                        # Comparar con tabla
                        st.markdown("---")
                        pedido = imagen_comparar.stem
                        if config['tabla'].exists():
                            tabla = cargar_tabla(str(config['tabla']))
                            datos_tabla = agrupar_por_pedido(tabla)
                            if pedido in datos_tabla:
                                info = datos_tabla[pedido]
                                st.subheader(f"📋 Datos correctos (tabla) para pedido {pedido}")
                                col1, col2, col3, col4 = st.columns(4)
                                col1.metric("Email", info['email'][:20] + "..." if len(info['email']) > 20 else info['email'])
                                col2.metric("Match", info['match'])
                                col3.metric("Cantidad", info['cantidad'])
                                col4.metric("Categoria", info['categoria'])

        # === TAB: Analisis Masivo OCR ===
        with debug_tab3:
            st.subheader("📊 Analisis Masivo - Ver donde falla OCR")
            st.info("Analiza todas las imagenes con EasyOCR para ver cuales necesitarian Claude Vision")

            if config['imagenes'].exists():
                imagenes = obtener_imagenes(config['imagenes'])
                num_imgs = len(imagenes)

                if num_imgs == 0:
                    st.warning("No hay imagenes en la carpeta")
                    max_analizar = 0
                elif num_imgs <= 20:
                    # Pocas imágenes, no mostrar slider
                    max_analizar = num_imgs
                    st.info(f"Se analizaran las {max_analizar} imagenes disponibles")
                else:
                    # Más de 20 imágenes, mostrar slider
                    max_analizar = st.slider("Cantidad de imagenes a analizar", 5, min(100, num_imgs), 20)

                if max_analizar > 0 and st.button(f"🔍 Analizar {max_analizar} imagenes con EasyOCR", type="primary"):
                    progress = st.progress(0)
                    status = st.empty()

                    resultados_analisis = []

                    for i, img in enumerate(imagenes[:max_analizar]):
                        status.text(f"Analizando {img.name}... ({i+1}/{max_analizar})")
                        progress.progress((i+1) / max_analizar)

                        res = extraer_datos_con_ocr(str(img))
                        valido, campos_ok, campos_fail = validar_resultado_ocr(res)

                        resultados_analisis.append({
                            'archivo': img.name,
                            'ocr_valido': '✅ Si' if valido else '❌ No',
                            'email': res.get('email') or '-',
                            'match': res.get('match') or '-',
                            'cantidad': res.get('cantidad') or '-',
                            'categoria': res.get('categoria') or '-',
                            'campos_ok': ', '.join(campos_ok) if campos_ok else '-',
                            'campos_faltantes': ', '.join(campos_fail) if campos_fail else '-',
                            'necesita_claude': 'Si' if not valido else 'No'
                        })

                    status.text("✅ Analisis completado")
                    progress.progress(1.0)

                    df_analisis = pd.DataFrame(resultados_analisis)

                    # Resumen
                    total = len(df_analisis)
                    validos = len(df_analisis[df_analisis['necesita_claude'] == 'No'])
                    fallback = total - validos

                    st.markdown("---")
                    col1, col2, col3 = st.columns(3)
                    col1.metric("✅ OCR Suficiente", validos)
                    col2.metric("❌ Necesita Claude", fallback)
                    col3.metric("📈 % Ahorro estimado", f"{(validos/total)*100:.0f}%")

                    # Filtrar
                    st.markdown("---")
                    filtro_resultado = st.radio(
                        "Filtrar resultados",
                        ["Todos", "Solo los que necesitan Claude", "Solo los que OCR funciona"],
                        horizontal=True
                    )

                    df_mostrar = df_analisis.copy()
                    if filtro_resultado == "Solo los que necesitan Claude":
                        df_mostrar = df_mostrar[df_mostrar['necesita_claude'] == 'Si']
                    elif filtro_resultado == "Solo los que OCR funciona":
                        df_mostrar = df_mostrar[df_mostrar['necesita_claude'] == 'No']

                    st.dataframe(df_mostrar, use_container_width=True)

                    # Descargar
                    csv = df_analisis.to_csv(index=False)
                    st.download_button(
                        "📥 Descargar analisis CSV",
                        data=csv,
                        file_name="analisis_ocr_easyocr.csv",
                        mime="text/csv"
                    )

        # === TAB: Cache ===
        with debug_tab4:
            st.subheader("💾 Cache de Resultados")
            cache = cargar_cache()

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Imagenes en cache", len(cache))
            with col2:
                cache_path = get_cache_path()
                if cache_path.exists():
                    size_kb = cache_path.stat().st_size / 1024
                    st.metric("Tamano cache", f"{size_kb:.1f} KB")
            with col3:
                if st.button("🗑️ Limpiar Cache"):
                    limpiar_cache()
                    st.success("Cache eliminado")
                    st.rerun()

            st.markdown("---")

            # Ver contenido del cache
            st.subheader("📋 Contenido del Cache")
            if cache:
                df_cache = pd.DataFrame([
                    {'hash': k[:8] + '...', **{key: v for key, v in val.items()}}
                    for k, val in list(cache.items())[:20]
                ])
                st.dataframe(df_cache, use_container_width=True)
                if len(cache) > 20:
                    st.info(f"Mostrando 20 de {len(cache)} entradas")
            else:
                st.info("El cache esta vacio")

    # === TAB VERIFICACION ===
    with tab_verificacion:
        email_usuario = get_usuario_email()
        carpeta_usuario = get_carpeta_usuario(email_usuario)
        config = cargar_config()

        if not config:
            st.error("Error cargando configuracion. Ve a la pestana Configuracion.")
            return

        # Mostrar info del usuario
        st.info(f"👤 Usuario: **{email_usuario or 'No autenticado'}**")

        # Verificar que hay archivos subidos
        carpeta_imagenes = carpeta_usuario / 'imagenes'
        carpeta_tabla = carpeta_usuario / 'tabla'

        # Buscar tabla existente
        tablas_existentes = []
        if carpeta_tabla.exists():
            tablas_existentes = list(carpeta_tabla.glob('*.csv')) + \
                               list(carpeta_tabla.glob('*.xlsx')) + \
                               list(carpeta_tabla.glob('*.xls'))

        if not tablas_existentes:
            st.warning("⚠️ No has subido una tabla de referencia.")
            st.write("Ve a la pestaña **Configuracion** y sube tu archivo CSV o Excel con los datos de referencia.")
            return

        # Usar la primera tabla encontrada
        tabla_path = tablas_existentes[0]

        # Obtener imagenes
        imagenes = []
        if carpeta_imagenes.exists():
            imagenes = obtener_imagenes(carpeta_imagenes)
        total_imagenes = len(imagenes)

        if total_imagenes == 0:
            st.warning("⚠️ No has subido imagenes de comprobantes.")
            st.write("Ve a la pestaña **Configuracion** y sube las imagenes de los comprobantes a verificar.")
            return

        # Actualizar config con rutas actuales
        config['imagenes'] = carpeta_imagenes
        config['tabla'] = tabla_path
        config['reporte'] = carpeta_usuario / 'reportes'

        st.info(f"📷 **{total_imagenes} imagenes** encontradas para verificar")

        # Opciones avanzadas
        with st.expander("⚙️ Opciones de procesamiento", expanded=True):
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("**🔍 Metodo de Extraccion**")
                metodo_extraccion = st.radio(
                    "Selecciona el metodo",
                    [
                        "🆓 OCR + Claude Fallback (Recomendado)",
                        "🆓 Solo OCR (Gratuito)",
                        "💰 Solo Claude Vision"
                    ],
                    help="OCR es gratuito pero menos preciso. Claude Vision es de pago pero muy preciso."
                )

                metodo_map = {
                    "🆓 OCR + Claude Fallback (Recomendado)": "ocr_fallback",
                    "🆓 Solo OCR (Gratuito)": "solo_ocr",
                    "💰 Solo Claude Vision": "solo_claude"
                }
                metodo = metodo_map[metodo_extraccion]

                if metodo == "solo_ocr":
                    st.success("💰 No se usara la API de Claude (ahorro total)")
                elif metodo == "ocr_fallback":
                    st.info("💡 Se usara Claude solo cuando EasyOCR falle")
                else:
                    st.warning("⚠️ Se usara Claude para todas las imagenes")

                # Solo EasyOCR disponible
                motor_ocr = "easyocr"

            with col2:
                usar_cache = st.checkbox("💾 Usar cache", value=True, help="Evita re-procesar imagenes ya analizadas")
                cache = cargar_cache()
                imagenes_en_cache = sum(1 for img in imagenes if calcular_hash_imagen(str(img)) in cache)
                if usar_cache and imagenes_en_cache > 0:
                    st.caption(f"📦 {imagenes_en_cache}/{total_imagenes} en cache")

                usar_paralelo = st.checkbox("⚡ Procesamiento paralelo", value=False, help="Mas rapido pero puede saturar")
                if usar_paralelo:
                    max_workers = st.slider("Workers", 2, 6, 3)
                else:
                    max_workers = 1

                tolerancia_email = st.slider(
                    "📧 Tolerancia email (%)",
                    80, 100, 90,
                    help="Porcentaje minimo de similitud para aceptar emails"
                ) / 100

        # Botones de accion
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            iniciar = st.button("🚀 Iniciar Verificacion", use_container_width=True, type="primary")

        # Boton para limpiar carpetas
        with st.expander("🧹 Limpiar carpetas de resultados"):
            carpeta_buenos = config['reporte'] / 'buenos'
            carpeta_regular = config['reporte'] / 'regular'
            carpeta_malos = config['reporte'] / 'malos'

            total_archivos = 0
            detalles = []
            for nombre, carpeta in [("buenos", carpeta_buenos), ("regular", carpeta_regular), ("malos", carpeta_malos)]:
                if carpeta.exists():
                    count = len(list(carpeta.iterdir()))
                    total_archivos += count
                    detalles.append(f"• {nombre}: {count} archivos")

            if total_archivos > 0:
                st.write("\n".join(detalles))
                st.write(f"**Total: {total_archivos} archivos**")

                if st.button("🗑️ Eliminar todos los archivos", type="secondary"):
                    eliminados = 0
                    for carpeta in [carpeta_buenos, carpeta_regular, carpeta_malos]:
                        if carpeta.exists():
                            for archivo in carpeta.iterdir():
                                try:
                                    archivo.unlink()
                                    eliminados += 1
                                except:
                                    pass
                    st.success(f"Se eliminaron {eliminados} archivos")
                    st.rerun()
            else:
                st.info("Las carpetas estan vacias")

        if iniciar:
            # Contenedores para actualizacion en tiempo real
            progress_bar = st.progress(0)
            status_text = st.empty()

            # Metricas en tiempo real
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            metric_ok = col1.empty()
            metric_parcial = col2.empty()
            metric_no_coincide = col3.empty()
            metric_no_encontrado = col4.empty()
            metric_ocr = col5.empty()
            metric_claude = col6.empty()

            # Log expandible
            log_container = st.expander("📋 Log de procesamiento", expanded=True)
            log_text = log_container.empty()
            log_messages = []

            # Inicializar contadores
            contadores = {
                'ok': 0, 'parcial': 0, 'no_coincide': 0, 'no_encontrado': 0,
                'ocr': 0, 'claude': 0, 'desde_cache': 0
            }

            def actualizar_metricas():
                metric_ok.metric("✅ OK", contadores['ok'])
                metric_parcial.metric("⚠️ Parcial", contadores['parcial'])
                metric_no_coincide.metric("❌ No coincide", contadores['no_coincide'])
                metric_no_encontrado.metric("🔍 No encontrado", contadores['no_encontrado'])
                metric_ocr.metric("🆓 OCR", contadores['ocr'])
                metric_claude.metric("💰 Claude", contadores['claude'])

            actualizar_metricas()

            # Inicializar cliente Claude (puede ser None si solo usamos OCR)
            client = None
            if metodo != "solo_ocr":
                status_text.text("🔌 Conectando con Claude Vision...")
                client = anthropic.Anthropic(api_key=config['api_key'])

            # Cargar tabla
            status_text.text("📊 Cargando tabla de referencia...")
            tabla = cargar_tabla(str(config['tabla']))
            datos_tabla = agrupar_por_pedido(tabla)

            # Cargar cache
            cache = cargar_cache()

            # Precarga del modelo OCR si es necesario
            if metodo != "solo_claude":
                status_text.text("🔧 Cargando EasyOCR (primera vez puede tardar)...")
                try:
                    get_easyocr_reader()
                except Exception as e:
                    st.error(f"Error cargando EasyOCR: {e}")
                    return

            resultados = []
            imagenes_ordenadas = sorted(imagenes)

            if usar_paralelo and max_workers > 1:
                # Procesamiento paralelo
                status_text.text(f"⚡ Procesando {total_imagenes} imagenes en paralelo ({max_workers} workers)...")

                args_list = [
                    (img, client, cache, usar_cache, datos_tabla, tolerancia_email, metodo, motor_ocr)
                    for img in imagenes_ordenadas
                ]

                completados = 0
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = {executor.submit(procesar_imagen_worker, args): args[0] for args in args_list}

                    for future in as_completed(futures):
                        resultado = future.result()
                        resultados.append(resultado)
                        completados += 1

                        # Actualizar contadores
                        if resultado['estado'] in ['OK', 'OK (EMAIL SIMILAR)']:
                            contadores['ok'] += 1
                        elif resultado['estado'] == 'PARCIAL':
                            contadores['parcial'] += 1
                        elif resultado['estado'] == 'NO ENCONTRADO EN TABLA':
                            contadores['no_encontrado'] += 1
                        else:
                            contadores['no_coincide'] += 1

                        # Contadores de metodo
                        metodo_usado = resultado.get('metodo_usado', '')
                        if 'OCR' in metodo_usado and 'fallback' not in metodo_usado.lower():
                            contadores['ocr'] += 1
                        elif 'Claude' in metodo_usado:
                            contadores['claude'] += 1

                        # Log
                        cache_tag = " [CACHE]" if resultado.get('desde_cache') else ""
                        metodo_tag = f" [{metodo_usado}]" if metodo_usado else ""
                        emoji = "✅" if resultado['estado'].startswith('OK') else ("⚠️" if resultado['estado'] == 'PARCIAL' else "❌")
                        log_messages.append(f"{emoji} {resultado['archivo']}{cache_tag}{metodo_tag}")

                        progress_bar.progress(completados / total_imagenes)
                        status_text.text(f"⚡ Procesando: {completados}/{total_imagenes}")
                        actualizar_metricas()
                        log_text.code('\n'.join(log_messages[-20:]))
            else:
                # Procesamiento secuencial
                for i, imagen in enumerate(imagenes_ordenadas):
                    progress_bar.progress((i + 1) / total_imagenes)
                    status_text.text(f"🔄 Procesando: {imagen.name} ({i+1}/{total_imagenes})")

                    args = (imagen, client, cache, usar_cache, datos_tabla, tolerancia_email, metodo, motor_ocr)
                    resultado = procesar_imagen_worker(args)
                    resultados.append(resultado)

                    # Actualizar contadores
                    if resultado['estado'] in ['OK', 'OK (EMAIL SIMILAR)']:
                        contadores['ok'] += 1
                    elif resultado['estado'] == 'PARCIAL':
                        contadores['parcial'] += 1
                    elif resultado['estado'] == 'NO ENCONTRADO EN TABLA':
                        contadores['no_encontrado'] += 1
                    else:
                        contadores['no_coincide'] += 1

                    # Contadores de metodo
                    metodo_usado = resultado.get('metodo_usado', '')
                    if 'OCR' in metodo_usado and 'fallback' not in metodo_usado.lower():
                        contadores['ocr'] += 1
                    elif 'Claude' in metodo_usado:
                        contadores['claude'] += 1

                    # Log
                    cache_tag = " [CACHE]" if resultado.get('desde_cache') else ""
                    metodo_tag = f" [{metodo_usado}]" if metodo_usado else ""
                    emoji = "✅" if resultado['estado'].startswith('OK') else ("⚠️" if resultado['estado'] == 'PARCIAL' else "❌")
                    log_messages.append(f"{emoji} {resultado['archivo']}{cache_tag}{metodo_tag}")

                    actualizar_metricas()
                    log_text.code('\n'.join(log_messages[-20:]))

            # Guardar cache actualizado
            guardar_cache(cache)

            # Ordenar resultados por archivo
            resultados.sort(key=lambda x: x['archivo'])

            # Guardar reporte CSV
            status_text.text("💾 Guardando reportes...")
            df = pd.DataFrame(resultados)

            # Seleccionar columnas para el reporte
            columnas_reporte = [
                'archivo', 'estado', 'pedido', 'metodo_usado',
                'email_imagen', 'email_tabla', 'email_ok', 'email_similar', 'similitud_email',
                'match_imagen', 'match_tabla', 'match_ok',
                'cantidad_imagen', 'cantidad_tabla', 'cantidad_ok',
                'categoria_imagen', 'categoria_tabla', 'categoria_ok',
                'desde_cache', 'fallback_usado'
            ]
            df_reporte = df[[c for c in columnas_reporte if c in df.columns]]

            reporte_path = config['reporte'] / 'reporte_verificacion.csv'
            df_reporte.to_csv(reporte_path, index=False)

            # Clasificar archivos
            status_text.text("📂 Clasificando archivos...")
            carpeta_buenos, carpeta_regular, carpeta_malos = clasificar_archivos(
                resultados, config['imagenes'], config['reporte']
            )

            # Completado
            progress_bar.progress(1.0)
            status_text.text("✅ ¡Verificacion completada!")

            st.success("🎉 Proceso completado")

            # Resumen final
            st.subheader("📊 Resumen Final")
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("✅ OK", contadores['ok'])
            col2.metric("⚠️ Parcial", contadores['parcial'])
            col3.metric("❌ No coincide", contadores['no_coincide'])
            col4.metric("🔍 No encontrado", contadores['no_encontrado'])

            # Estadisticas de ahorro
            st.subheader("💰 Ahorro en API")
            col1, col2, col3 = st.columns(3)
            col1.metric("🆓 Procesadas con OCR", contadores['ocr'])
            col2.metric("💰 Procesadas con Claude", contadores['claude'])

            if total_imagenes > 0:
                porcentaje_ahorro = (contadores['ocr'] / total_imagenes) * 100
                col3.metric("📈 Ahorro estimado", f"{porcentaje_ahorro:.0f}%")

            # Mostrar rutas
            st.subheader("📁 Archivos clasificados")
            st.write(f"- **Buenos:** {carpeta_buenos}")
            st.write(f"- **Regular:** {carpeta_regular}")
            st.write(f"- **Malos:** {carpeta_malos}")
            st.write(f"- **Reporte:** {reporte_path}")

            # Mostrar tabla de resultados
            st.subheader("📋 Detalle de resultados")

            # Filtros
            col1, col2 = st.columns(2)
            with col1:
                filtro_estado = st.selectbox(
                    "Filtrar por estado:",
                    ["Todos", "OK", "OK (EMAIL SIMILAR)", "PARCIAL", "NO COINCIDE", "NO ENCONTRADO EN TABLA"]
                )
            with col2:
                filtro_metodo = st.selectbox(
                    "Filtrar por metodo:",
                    ["Todos", "OCR", "Claude Vision", "Claude Vision (fallback)"]
                )

            df_mostrar = df_reporte.copy()
            if filtro_estado != "Todos":
                df_mostrar = df_mostrar[df_mostrar['estado'] == filtro_estado]
            if filtro_metodo != "Todos":
                df_mostrar = df_mostrar[df_mostrar['metodo_usado'].str.contains(filtro_metodo.split()[0], case=False, na=False)]

            st.dataframe(df_mostrar, use_container_width=True)

            # Botones de descarga
            st.subheader("📥 Descargar Reportes")
            col1, col2 = st.columns(2)

            with col1:
                csv = df_reporte.to_csv(index=False)
                st.download_button(
                    label="📥 Descargar CSV",
                    data=csv,
                    file_name="reporte_verificacion.csv",
                    mime="text/csv"
                )

            with col2:
                excel_data = exportar_excel(df_reporte)
                st.download_button(
                    label="📥 Descargar Excel",
                    data=excel_data,
                    file_name="reporte_verificacion.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )


# Para ejecutar directamente (debug):
# if __name__ == "__main__":
#     render()
